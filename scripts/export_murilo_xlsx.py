"""
Demanda 9 (higiene visual) -- Gera uma versao .xlsx AMIGAVEL do pacote
for_murilo, para Murilo preencher com conforto.

Lê:
  reports/tail_pilot_120_for_murilo_2026-04-10.csv

Gera:
  reports/tail_pilot_120_for_murilo_2026-04-10.xlsx

Formatacao:
  - header row fixo (freeze_panes)
  - autofilter no header
  - colunas auto-largas (cap por coluna)
  - colunas murilo_* destacadas (verde claro) -- sao as unicas editaveis
  - colunas r1_* destacadas (azul claro) -- sao contexto
  - flags cruas (wine_filter_category, y2_any_not_wine_or_spirit, etc.) destacadas (amarelo claro)
  - data validation (dropdown) nas 4 colunas categoricas do Murilo:
      murilo_business_class ∈ {MATCH_RENDER, MATCH_IMPORT, STANDALONE_WINE, NOT_WINE}
      murilo_review_state   ∈ {RESOLVED, SECOND_REVIEW, UNRESOLVED}
      murilo_confidence     ∈ {HIGH, MEDIUM, LOW}
      murilo_action         ∈ {ALIAS, IMPORT_THEN_ALIAS, KEEP_STANDALONE, SUPPRESS}
  - wrap text nas colunas longas (nome, top3_*_summary, r1_reason_long)
  - linhas com altura padrao

Tambem salva um CSV cópia com BOM (utf-8-sig) para Excel reconhecer UTF-8
em usuarios que preferirem CSV:
  reports/tail_pilot_120_for_murilo_2026-04-10_excel.csv

Uso:
  python scripts/export_murilo_xlsx.py
"""

import csv
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


REPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "reports")
IN_CSV = os.path.join(REPORT_DIR, "tail_pilot_120_for_murilo_2026-04-10.csv")
OUT_XLSX = os.path.join(REPORT_DIR, "tail_pilot_120_for_murilo_2026-04-10.xlsx")
OUT_CSV_BOM = os.path.join(REPORT_DIR, "tail_pilot_120_for_murilo_2026-04-10_excel.csv")


# Classificacao das colunas
MURILO_EDIT_COLS = {
    "murilo_business_class",
    "murilo_review_state",
    "murilo_confidence",
    "murilo_action",
    "murilo_notes",
}
R1_CTX_COLS = {
    "r1_business_class", "r1_review_state", "r1_confidence",
    "r1_action", "r1_data_quality", "r1_product_impact",
    "r1_recommended_universe", "r1_recommended_candidate_id",
    "r1_reason_short", "r1_reason_long", "r1_match_blockers",
    "murilo_review_reason",
}
RAW_FLAG_COLS = {
    "wine_filter_category", "y2_any_not_wine_or_spirit",
    "y2_status_set", "y2_present", "no_source_flag",
    "reason_short_proxy", "block", "overflow_from",
    "pilot_bucket_proxy",
}

# Largura aproximada (caracteres Excel) por coluna
COL_WIDTH = {
    "render_wine_id": 11,
    "pilot_bucket_proxy": 28,
    "block": 22,
    "overflow_from": 14,
    "nome": 45,
    "produtor": 30,
    "safra": 7,
    "tipo": 12,
    "preco_min": 10,
    "wine_sources_count_live": 7,
    "stores_count_live": 7,
    "no_source_flag": 7,
    "y2_present": 7,
    "y2_status_set": 14,
    "y2_any_not_wine_or_spirit": 9,
    "wine_filter_category": 16,
    "reason_short_proxy": 32,
    "top1_render_candidate_id": 11,
    "top1_render_channel": 20,
    "top1_render_score": 9,
    "top1_render_gap": 9,
    "top1_import_candidate_id": 11,
    "top1_import_channel": 20,
    "top1_import_score": 9,
    "top1_import_gap": 9,
    "best_overall_universe": 12,
    "best_overall_channel": 20,
    "best_overall_score": 9,
    "top1_render_human": 55,
    "top1_import_human": 55,
    "top3_render_summary": 80,
    "top3_import_summary": 80,
    "r1_business_class": 17,
    "r1_review_state": 16,
    "r1_confidence": 12,
    "r1_action": 18,
    "r1_data_quality": 12,
    "r1_product_impact": 12,
    "r1_recommended_universe": 12,
    "r1_recommended_candidate_id": 11,
    "r1_reason_short": 40,
    "r1_reason_long": 70,
    "r1_match_blockers": 30,
    "murilo_review_reason": 40,
    "murilo_business_class": 18,
    "murilo_review_state": 17,
    "murilo_confidence": 13,
    "murilo_action": 20,
    "murilo_notes": 50,
}
WRAP_COLS = {
    "nome", "produtor",
    "top1_render_human", "top1_import_human",
    "top3_render_summary", "top3_import_summary",
    "r1_reason_short", "r1_reason_long", "r1_match_blockers",
    "reason_short_proxy", "murilo_review_reason", "murilo_notes",
}

# Cores
GREEN_FILL = PatternFill("solid", fgColor="DFF0D8")   # murilo edit cells
BLUE_FILL = PatternFill("solid", fgColor="D9E8F5")    # r1 context cells
YELLOW_FILL = PatternFill("solid", fgColor="FFF3CD")  # raw flags
HEADER_FILL_DEFAULT = PatternFill("solid", fgColor="343A40")
HEADER_FILL_MURILO = PatternFill("solid", fgColor="198754")
HEADER_FILL_R1 = PatternFill("solid", fgColor="0D6EFD")
HEADER_FILL_FLAG = PatternFill("solid", fgColor="FFC107")

HEADER_FONT_LIGHT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FONT_DARK = Font(bold=True, color="212529", size=10)

THIN = Side(border_style="thin", color="D3D3D3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def main():
    # Read input CSV (UTF-8 sem BOM)
    with open(IN_CSV, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames
        rows = list(reader)
    print(f"[read] {len(rows)} rows, {len(headers)} cols")

    # ---------- XLSX ----------
    wb = Workbook()
    ws = wb.active
    ws.title = "pilot_120"

    # Header
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT_LIGHT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if h in MURILO_EDIT_COLS:
            cell.fill = HEADER_FILL_MURILO
        elif h in R1_CTX_COLS:
            cell.fill = HEADER_FILL_R1
        elif h in RAW_FLAG_COLS:
            cell.fill = HEADER_FILL_FLAG
            cell.font = HEADER_FONT_DARK
        else:
            cell.fill = HEADER_FILL_DEFAULT
        cell.border = BORDER

    # Rows
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(headers, 1):
            v = row.get(h, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            align = Alignment(
                horizontal="left" if h in WRAP_COLS or h in MURILO_EDIT_COLS else "left",
                vertical="top",
                wrap_text=(h in WRAP_COLS),
            )
            cell.alignment = align
            if h in MURILO_EDIT_COLS:
                cell.fill = GREEN_FILL
            elif h in R1_CTX_COLS:
                cell.fill = BLUE_FILL
            elif h in RAW_FLAG_COLS:
                cell.fill = YELLOW_FILL
            cell.border = BORDER

    # Column widths
    for c_idx, h in enumerate(headers, 1):
        letter = get_column_letter(c_idx)
        width = COL_WIDTH.get(h, 18)
        ws.column_dimensions[letter].width = width

    # Row heights: cabeca 38, linhas normais 70 (bastante para wrap de top3_* e reason_long)
    ws.row_dimensions[1].height = 38
    for r_idx in range(2, len(rows) + 2):
        ws.row_dimensions[r_idx].height = 70

    # Freeze header
    ws.freeze_panes = "A2"

    # AutoFilter
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    # ---------- Data validation nos 4 campos ----------
    def add_dropdown(col_name, values):
        if col_name not in headers:
            return
        c_idx = headers.index(col_name) + 1
        letter = get_column_letter(c_idx)
        # CSV list format para Excel: "VAL1,VAL2,VAL3"; precisa entre aspas duplas se tiver separadores
        formula = '"' + ",".join(values) + '"'
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=False,  # False = mostra seta (convencao invertida openpyxl)
            showErrorMessage=True,
            errorTitle="Valor invalido",
            error=f"Use um dos valores: {', '.join(values)}",
        )
        dv.add(f"{letter}2:{letter}{len(rows) + 1}")
        ws.add_data_validation(dv)

    add_dropdown("murilo_business_class",
                 ["MATCH_RENDER", "MATCH_IMPORT", "STANDALONE_WINE", "NOT_WINE"])
    add_dropdown("murilo_review_state",
                 ["RESOLVED", "SECOND_REVIEW", "UNRESOLVED"])
    add_dropdown("murilo_confidence",
                 ["HIGH", "MEDIUM", "LOW"])
    add_dropdown("murilo_action",
                 ["ALIAS", "IMPORT_THEN_ALIAS", "KEEP_STANDALONE", "SUPPRESS"])

    # ---------- Sheet de LEGENDA/INSTRUCOES ----------
    ws2 = wb.create_sheet("leia_primeiro")
    ws2.column_dimensions["A"].width = 110
    lines = [
        ("Pilot 120 -- Instrucoes rapidas para Murilo", HEADER_FONT_LIGHT, HEADER_FILL_DEFAULT),
        ("", None, None),
        ("Preencha APENAS as 5 colunas em VERDE (murilo_*).", None, GREEN_FILL),
        ("As colunas em AMARELO sao flags cruas operacionais (contexto).", None, YELLOW_FILL),
        ("As colunas em AZUL sao a classificacao R1 do Claude (referencia).", None, BLUE_FILL),
        ("", None, None),
        ("Colunas obrigatorias para preencher:", Font(bold=True), None),
        ("  murilo_business_class  -> MATCH_RENDER / MATCH_IMPORT / STANDALONE_WINE / NOT_WINE", None, None),
        ("  murilo_review_state    -> RESOLVED / SECOND_REVIEW / UNRESOLVED", None, None),
        ("  murilo_confidence      -> HIGH / MEDIUM / LOW", None, None),
        ("  murilo_action          -> ALIAS / IMPORT_THEN_ALIAS / KEEP_STANDALONE / SUPPRESS", None, None),
        ("  murilo_notes           -> texto livre (opcional mas util)", None, None),
        ("", None, None),
        ("Regras sagradas:", Font(bold=True), None),
        ("  1. NAO renomeie o arquivo nem a aba 'pilot_120'.", None, None),
        ("  2. NAO altere colunas fora das cinco murilo_*.", None, None),
        ("  3. NAO apague linhas.", None, None),
        ("  4. UNRESOLVED vai em review_state, nunca em business_class.", None, None),
        ("  5. y2_any_not_wine_or_spirit e baseline, nao verdade.", None, None),
        ("", None, None),
        ("Quando terminar, salve o arquivo e exporte de volta para CSV", Font(bold=True), None),
        ("(Salvar como -> CSV UTF-8), mantendo o nome:", None, None),
        ("tail_pilot_120_for_murilo_2026-04-10.csv", Font(name="Consolas", bold=True), None),
        ("", None, None),
        ("Depois rode a sequencia:", Font(bold=True), None),
        ("  python scripts/validate_murilo_csv.py", Font(name="Consolas"), None),
        ("  python scripts/compare_claude_vs_murilo.py", Font(name="Consolas"), None),
        ("  python scripts/build_adjudication_template.py", Font(name="Consolas"), None),
        ("", None, None),
        ("Relacao canonica business_class -> action:", Font(bold=True), None),
        ("  MATCH_RENDER    -> ALIAS", None, None),
        ("  MATCH_IMPORT    -> IMPORT_THEN_ALIAS", None, None),
        ("  STANDALONE_WINE -> KEEP_STANDALONE", None, None),
        ("  NOT_WINE        -> SUPPRESS", None, None),
        ("", None, None),
        ("(ver reports/tail_pilot_120_murilo_instructions_2026-04-10.md para texto completo)", None, None),
    ]
    for i, (text, font, fill) in enumerate(lines, 1):
        c = ws2.cell(row=i, column=1, value=text)
        if font:
            c.font = font
        if fill:
            c.fill = fill
        c.alignment = Alignment(vertical="center", wrap_text=False)

    # Mover a aba de leia_primeiro para o inicio
    wb.move_sheet(ws2, offset=-1)

    # Salvar
    wb.save(OUT_XLSX)
    print(f"[write] {OUT_XLSX}")

    # ---------- CSV com BOM para Excel ----------
    with open(OUT_CSV_BOM, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        w.writerows(rows)
    print(f"[write] {OUT_CSV_BOM}  (CSV com BOM para Excel)")

    print()
    print("=== exports prontos ===")
    print(f"  {OUT_XLSX}")
    print(f"  {OUT_CSV_BOM}")


if __name__ == "__main__":
    main()
